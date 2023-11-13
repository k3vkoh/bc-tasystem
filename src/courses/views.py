from django.views import View
from django.shortcuts import render, redirect
from .models import Course
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from openpyxl import load_workbook
from users.models import CustomUser as User
from random import randint
from django.views.generic import ListView, DetailView
from django.contrib import messages
from users.instructor_data import instructors

from random import randint


class UploadView(LoginRequiredMixin, UserPassesTestMixin, View):
    def get(self, request):
        return render(request, 'upload.html')

    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            self.process_excel_file(excel_file)
        messages.success(self.request, 'Successfully Uploaded Excel File')
        return render(request, 'upload.html')

    def process_excel_file(self, excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            if not row[0] or row[0] == 'Term':
                continue
            instructor = self.get_or_create_instructor(row)
            self.create_course(row, instructor)

    def get_email(self, first_name, last_name):
        return f'{last_name}@bc.edu' if f'{last_name}, {first_name}' not in instructors else instructors[f'{last_name}, {first_name}']['Short Email']

    def get_or_create_instructor(self, row):
        instructor_first_name = row[6].split(',')[1].strip()
        instructor_last_name = row[6].split(',')[0].strip()

        email = self.get_email(instructor_first_name, instructor_last_name)

        # FOR TESTING PURPOSES ONLY
        if randint(0, 5) == 1:
            instructor_first_name = 'Andy'
            instructor_last_name = 'Deng'
            email = 'denga@bc.edu'

        instructor, created = User.objects.get_or_create(
            first_name=instructor_first_name,
            last_name=instructor_last_name,
            defaults={
                'professor': True,
                'email': email,
                'eagleid': self.generate_eagleid()
            }
        )

        if created:
            instructor.set_password('password')
            instructor.save()

        return instructor

    def create_course(self, row, instructor):
        excluded_lectures = ["Computer Science I",
                             "Computer Science II", "Computer Organization and Lab"]

        if row[5] in excluded_lectures and row[1] == "Lecture":
            return

        num_tas = 1 if row[1] == "Discussion" or row[1] == "Lab" else int(
            row[9]) // 20

        num_tas = 1 if num_tas == 0 else num_tas

        new_class = Course.objects.create(
            term=row[0],
            class_type=row[1],
            course=row[3],
            section=row[4],
            course_title=row[5],
            instructor_first_name=instructor.first_name,
            instructor_last_name=instructor.last_name,
            room_name=row[7],
            timeslot=row[8],
            max_enroll=row[9],
            room_size=row[10],
            num_tas=num_tas,
            description=row[12],
            professor=instructor
        )
        instructor.courses.add(new_class)

    def generate_eagleid(self):
        eagleid = randint(10000000, 99999999)
        while User.objects.filter(eagleid=eagleid).exists():
            eagleid = randint(10000000, 99999999)
        return eagleid

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser


class CloseView(LoginRequiredMixin, UserPassesTestMixin, View):
    def post(self, request):
        self.archive_courses()
        messages.success(self.request, 'Successfully Closed Courses')
        return redirect('courses:manage-course')

    def archive_courses(self):
        current_courses = Course.objects.all()
        for course in current_courses:
            course.is_active = False
            course.save()

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser


class ListView(LoginRequiredMixin, ListView):
    model = Course
    template_name = 'list.html'
    ordering = ['course']
    context_object_name = 'course_data'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['professors'] = User.objects.filter(professor=True)
        return context

    def get_queryset(self):
        user = self.request.user
        professor_id = self.request.GET.get('professor_id', None)

        if professor_id:
            return Course.objects.filter(professor__id=professor_id, is_active=True)

        if user.is_student() or user.is_superuser:
            return Course.objects.filter(is_active=True)

        return Course.objects.filter(professor=user,is_active=True)


class CourseDetailView(LoginRequiredMixin, DetailView):
    model = Course
    template_name = 'course_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['course'] = Course.objects.get(pk=self.kwargs.get('pk'))
        context['is_professor'] = self.is_professor()
        context['at_max_apps'] = self.at_max_apps()
        context['has_applied'] = self.has_applied()
        context['is_ta'] = self.is_ta()
        return context

    def is_professor(self):
        return self.request.user.is_professor()

    def at_max_apps(self):
        return self.request.user.reached_max_applications()

    def has_applied(self):
        return self.request.user.already_applied_to_course(self.get_object())

    def is_ta(self):
        return self.request.user.is_ta()
